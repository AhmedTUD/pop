#!/usr/bin/env python3
"""
Enhanced Excel export with local images support and better formatting
"""

import os
import tempfile
import time
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# إعدادات محسنة للصور
PYTHONANYWHERE_SETTINGS = {
    'MAX_RETRIES': 3,
    'TIMEOUT': 30,
    'DELAY_BETWEEN_IMAGES': 1.5,
    'MAX_IMAGE_SIZE_MB': 15,
    'IMAGE_QUALITY': 85,
    'MAX_IMAGE_DIMENSIONS': (800, 600)
}

from datetime import datetime

# استيراد دوال الوقت المحلي
try:
    from app import get_local_time_string, LOCAL_TIMEZONE

except ImportError:
    try:
        # محاولة استيراد إعدادات المنطقة الزمنية
        from timezone_config import get_timezone_from_env
        LOCAL_TIMEZONE = get_timezone_from_env()
        
        def get_local_time_string():
            from datetime import datetime
            return datetime.now(LOCAL_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')
        

    except ImportError:
        # الحل الافتراضي مع المنطقة الزمنية المصرية
        from datetime import datetime, timezone, timedelta
        try:
            from zoneinfo import ZoneInfo
            LOCAL_TIMEZONE = ZoneInfo("Africa/Cairo")
        except ImportError:
            try:
                import pytz
                LOCAL_TIMEZONE = pytz.timezone("Africa/Cairo")
            except ImportError:
                # Final fallback - Egypt winter time (UTC+2)
                LOCAL_TIMEZONE = timezone(timedelta(hours=2))
        
        def get_local_time_string():
            return datetime.now(LOCAL_TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')
        


def process_multiple_images(image_urls, max_images=None):
    """
    معالجة صور متعددة مع تحسينات خاصة بـ PythonAnywhere
    
    Args:
        image_urls: قائمة بروابط الصور
        max_images: الحد الأقصى لعدد الصور
    
    Returns:
        list: قائمة بالصور المحملة بنجاح
    """
    if not image_urls:
        return []
    
    processed_images = []
    failed_count = 0
    
    # تحديد عدد الصور للمعالجة
    if max_images is None:
        max_images = len(image_urls)
    else:
        max_images = min(max_images, len(image_urls))
    
    print(f"🖼️ بدء معالجة {len(image_urls)} صورة (سيتم معالجة {max_images})")
    
    for i, image_url in enumerate(image_urls[:max_images]):
        print(f"\n📸 معالجة الصورة {i+1}/{min(len(image_urls), max_images)}")
        print(f"   الرابط: {image_url[:60]}...")
        
        # تأخير بين الصور لتجنب مشاكل الشبكة
        if i > 0:
            delay = PYTHONANYWHERE_SETTINGS['DELAY_BETWEEN_IMAGES']
            print(f"   ⏳ انتظار {delay} ثانية...")
            time.sleep(delay)
        
        try:
            # تحميل الصورة (محلية فقط)
            img_buffer = load_local_image(image_url)
            
            if img_buffer:
                processed_images.append({
                    'buffer': img_buffer,
                    'url': image_url,
                    'index': i
                })
                print(f"   ✅ تم تحميل الصورة بنجاح")
            else:
                failed_count += 1
                print(f"   ❌ فشل في تحميل الصورة")
                
        except Exception as e:
            failed_count += 1
            print(f"   ❌ خطأ في معالجة الصورة: {e}")
    
    success_rate = len(processed_images) / len(image_urls[:max_images]) * 100
    print(f"\n📊 نتيجة المعالجة:")
    print(f"   نجح: {len(processed_images)}")
    print(f"   فشل: {failed_count}")
    print(f"   معدل النجاح: {success_rate:.1f}%")
    
    return processed_images

def load_local_image(image_path):
    """
    تحميل صورة محلية بدون أي معالجة - الملف الأصلي كما هو
    
    Args:
        image_path: مسار الصورة المحلية
    
    Returns:
        BytesIO: الصورة الأصلية بدون معالجة أو None في حالة الخطأ
    """
    try:
        # محاولة قراءة الملف المحلي
        local_path = os.path.join('static/uploads', image_path)
        if not os.path.exists(local_path):
            print(f"   ❌ الملف المحلي غير موجود: {local_path}")
            return None
        
        # قراءة الملف الأصلي بدون أي معالجة
        with open(local_path, 'rb') as f:
            original_img_data = f.read()
        
        if len(original_img_data) < 1000:  # أقل من 1 كيلوبايت
            print(f"   ⚠️ الملف صغير جداً: {len(original_img_data)} bytes")
            return None
        
        # التحقق من صحة الصورة فقط (بدون معالجة)
        try:
            img_test = PILImage.open(BytesIO(original_img_data))
            img_test.verify()  # التحقق من صحة الصورة فقط
        except Exception as verify_error:
            print(f"   ❌ الصورة تالفة: {verify_error}")
            return None
        
        # إرجاع الصورة الأصلية بدون أي تعديل
        img_buffer = BytesIO(original_img_data)
        img_buffer.seek(0)
        
        print(f"   ✅ تم تحميل الصورة الأصلية بدون معالجة ({len(original_img_data)} bytes)")
        return img_buffer
        
    except Exception as e:
        print(f"   ❌ خطأ في تحميل الصورة المحلية: {e}")
        return None



def create_enhanced_excel_with_images(data_entries, filename):
    """
    إنشاء ملف Excel محسن مع الصور والتنسيق المحسن
    
    Args:
        data_entries: بيانات الإدخالات
        filename: اسم الملف
    
    Returns:
        str: مسار الملف المؤقت أو None في حالة الخطأ
    """
    try:
        # طباعة معلومات البيئة
        print("🌐 استخدام إعدادات محسنة للصور المحلية")
        
        print(f"📊 بدء إنشاء تقرير Excel مع {len(data_entries)} إدخال")
        
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "POP Materials Report"
        
        # تعريف الألوان والأنماط
        colors = {
            'header': 'FF366092',
            'alt_row': 'FFF2F2F2',
            'border': 'FF000000',
            'text': 'FF000000',
            'white': 'FFFFFFFF'
        }
        
        # تعريف الخطوط المحسنة (أكثر وضوحاً)
        header_font = Font(name='Calibri', size=14, bold=True, color=colors['white'])
        data_font = Font(name='Calibri', size=11, color=colors['text'])
        
        # تعريف الحدود
        thin_border = Border(
            left=Side(style='thin', color=colors['border']),
            right=Side(style='thin', color=colors['border']),
            top=Side(style='thin', color=colors['border']),
            bottom=Side(style='thin', color=colors['border'])
        )
        
        # العناوين بدون ID
        headers = [
            'Employee Name', 'Employee Code', 'Branch', 'Shop Code', 
            'Model', 'Display Type', 'Selected Materials', 'Missing Materials', 
            'Comments', 'Images Count', 'Date', 'Image Preview'
        ]
        
        # إضافة العناوين
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # سيتم تعيين عرض الأعمدة تلقائياً في النهاية بناءً على المحتوى
        
        # إضافة البيانات
        current_row = 2
        
        for entry in data_entries:
            # تحديد لون الصف (متناوب)
            row_fill = PatternFill(
                start_color=colors['alt_row'] if current_row % 2 == 0 else colors['white'],
                end_color=colors['alt_row'] if current_row % 2 == 0 else colors['white'],
                fill_type='solid'
            )
            
            # البيانات الأساسية (بدون ID)
            # الترتيب في الاستعلام: images(9), date(10), comment(11)
            entry_data = [
                entry[1],  # Employee Name
                entry[2],  # Employee Code
                entry[3],  # Branch
                entry[4] if entry[4] else 'N/A',  # Shop Code
                entry[5],  # Model
                entry[6],  # Display Type
                entry[7].replace(',', '\n') if entry[7] else 'None',  # Selected Materials
                entry[8].replace(',', '\n') if entry[8] else 'None',  # Missing Materials
                entry[11] if entry[11] else 'No comment',  # Comments
                len(entry[9].split(',')) if entry[9] else 0,  # Images Count
                entry[10]  # Date
            ]
            
            # إضافة البيانات إلى الخلايا
            for col, value in enumerate(entry_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', 
                                         vertical='top', wrap_text=True)
            
            # تعيين ارتفاع الصف لاستيعاب النص متعدد الأسطر والصور
            # حساب عدد الأسطر المطلوبة للمواد (بعد التحويل إلى أسطر منفصلة)
            selected_lines = len(entry[7].replace(',', '\n').split('\n')) if entry[7] else 1
            missing_lines = len(entry[8].replace(',', '\n').split('\n')) if entry[8] else 1
            max_material_lines = max(selected_lines, missing_lines)
            
            # معالجة الصور
            images_data = entry[9] if entry[9] else ''
            if images_data:
                # تنظيف وتصفية روابط الصور
                raw_urls = [url.strip() for url in images_data.split(',') if url.strip()]
                image_urls = [url for url in raw_urls if url and len(url) > 10]  # تصفية الروابط الفارغة أو القصيرة
                
                print(f"\n🖼️ الصف {current_row}: معالجة {len(image_urls)} صورة")
                
                # تعيين ارتفاع الصف بناءً على المحتوى والصور
                min_height_for_materials = max_material_lines * 15 + 10
                # حساب الارتفاع المطلوب للصور (أكبر صورة + مساحة إضافية)
                max_image_height = 150  # الحد الأقصى لارتفاع الصور
                required_height = max(min_height_for_materials, max_image_height + 20)
                ws.row_dimensions[current_row].height = required_height
                
                # معالجة الصور باستخدام الدالة المحسنة - جميع الصور بدون حد أقصى
                processed_images = process_multiple_images(image_urls, max_images=len(image_urls))
                
                # إدراج الصور في Excel
                images_added = 0
                for img_data in processed_images:
                    try:
                        img_buffer = img_data['buffer']
                        img_index = img_data['index']
                        
                        # إنشاء صورة Excel من الملف الأصلي
                        excel_img = ExcelImage(img_buffer)
                        
                        # الحصول على الأبعاد الأصلية
                        original_width = excel_img.width
                        original_height = excel_img.height
                        
                        # استخدام الأبعاد الأصلية مع حد أقصى معقول للعرض في Excel
                        # إذا كانت الصورة كبيرة جداً، نحدد حد أقصى للعرض مع الحفاظ على النسبة
                        max_display_width = 200  # حد أقصى للعرض في Excel
                        max_display_height = 150  # حد أقصى للارتفاع في Excel
                        
                        # تحديد ما إذا كنا بحاجة لتقليل الحجم للعرض في Excel فقط
                        if original_width > max_display_width or original_height > max_display_height:
                            # حساب النسبة للحفاظ على التناسب
                            width_ratio = max_display_width / original_width
                            height_ratio = max_display_height / original_height
                            scale_ratio = min(width_ratio, height_ratio)
                            
                            excel_img.width = int(original_width * scale_ratio)
                            excel_img.height = int(original_height * scale_ratio)
                        # إذا كانت الصورة صغيرة، نتركها كما هي
                        # لا نقوم بتكبيرها لتجنب فقدان الجودة
                        
                        # تحديد موقع الصورة - بدءاً من العمود 12 (L) وما بعده
                        image_column = 12 + img_index  # العمود L, M, N, O, P, Q... للصور
                        col_letter = get_column_letter(image_column)
                        
                        # تعيين عرض العمود للصور
                        ws.column_dimensions[col_letter].width = 25
                        
                        # حساب الموقع المتوسط للصورة في الخلية (توسيط)
                        cell_width_pixels = 25 * 7  # تقريبي: عرض العمود بالبكسل
                        cell_height_pixels = required_height * 1.33  # تقريبي: ارتفاع الصف بالبكسل
                        
                        # حساب الإزاحة للتوسيط
                        offset_x = max(0, (cell_width_pixels - excel_img.width) // 2)
                        offset_y = max(0, (cell_height_pixels - excel_img.height) // 2)
                        
                        # تعيين موقع الصورة مع التوسيط
                        excel_img.anchor = f"{col_letter}{current_row}"
                        
                        # إضافة الصورة
                        ws.add_image(excel_img)
                        images_added += 1
                        
                        print(f"   ✅ تم إدراج الصورة {img_index + 1} في العمود {col_letter}{current_row}")
                        
                    except Exception as e:
                        print(f"   ❌ خطأ في إدراج الصورة {img_index + 1}: {e}")
                
                # إضافة نص في خلية الصور
                failed_count = len(image_urls) - images_added
                if images_added > 0:
                    img_text = f"{images_added} of {len(image_urls)} images loaded"
                    if failed_count > 0:
                        img_text += f" ({failed_count} failed)"
                else:
                    img_text = f"0 of {len(image_urls)} images (all failed)"
                
                print(f"   📊 النتيجة النهائية: {img_text}")
                
                img_cell = ws.cell(row=current_row, column=12, value=img_text)
                img_cell.font = data_font
                img_cell.fill = row_fill
                img_cell.border = thin_border
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # لا توجد صور - تعيين ارتفاع الصف للنص فقط (مقلل)
                min_height_for_materials = max_material_lines * 15 + 5
                ws.row_dimensions[current_row].height = max(35, min_height_for_materials)
                
                img_cell = ws.cell(row=current_row, column=12, value="No images")
                img_cell.font = data_font
                img_cell.fill = row_fill
                img_cell.border = thin_border
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # إضافة ملخص في النهاية
        summary_row = current_row + 2
        
        # عنوان الملخص
        summary_cell = ws.cell(row=summary_row, column=1, value="Report Summary")
        summary_cell.font = Font(name='Arial', size=14, bold=True, color=colors['white'])
        summary_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
        ws.merge_cells(f'A{summary_row}:L{summary_row}')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # بيانات الملخص
        total_entries = len(data_entries)
        total_images = sum(len(entry[9].split(',')) if entry[9] else 0 for entry in data_entries)
        unique_employees = len(set(entry[2] for entry in data_entries))
        unique_branches = len(set(entry[3] for entry in data_entries))
        
        summary_data = [
            f"Total Entries: {total_entries}",
            f"Total Images: {total_images}",
            f"Unique Employees: {unique_employees}",
            f"Unique Branches: {unique_branches}",
            f"Report Date: {get_local_time_string()}"
        ]
        
        for i, summary_text in enumerate(summary_data):
            cell = ws.cell(row=summary_row + 1 + i, column=1, value=summary_text)
            cell.font = Font(name='Arial', size=10, bold=True)
            ws.merge_cells(f'A{summary_row + 1 + i}:L{summary_row + 1 + i}')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # الحل النهائي: استخدام auto_fit للأعمدة
        # تعيين عرض مناسب لكل عمود بناءً على المحتوى
        
        # أعرض محسنة للأعمدة بناءً على المحتوى
        column_widths = {
            1: 22,   # Employee Name - زيادة العرض
            2: 16,   # Employee Code  
            3: 28,   # Branch - زيادة العرض
            4: 14,   # Shop Code
            5: 30,   # Model - زيادة العرض
            6: 22,   # Display Type
            7: 38,   # Selected Materials - زيادة العرض
            8: 38,   # Missing Materials - زيادة العرض
            9: 25,   # Comments - زيادة العرض
            10: 14,  # Images Count
            11: 20,  # Date - زيادة العرض
            12: 25,  # Image Preview/Image 1
        }
        
        # إضافة أعمدة إضافية للصور (حتى 10 صور)
        for i in range(13, 22):  # أعمدة M إلى U للصور الإضافية
            column_widths[i] = 25
        
        # تطبيق الأعرض
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = width
        
        # حفظ الملف
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        wb.save(temp_path)
        
        return temp_path
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {e}")
        return None



def calculate_optimal_row_height(cell_values, font_size=11):
    """
    حساب الارتفاع الأمثل للصف بناءً على محتوى الخلايا
    
    Args:
        cell_values: قائمة بقيم الخلايا في الصف
        font_size: حجم الخط
    
    Returns:
        float: الارتفاع المطلوب بالنقاط
    """
    max_lines = 1
    max_chars_per_line = 0
    
    for value in cell_values:
        if value and isinstance(value, str):
            # حساب عدد الأسطر الفعلية
            lines = value.split('\n')
            max_lines = max(max_lines, len(lines))
            
            # حساب أطول سطر
            for line in lines:
                max_chars_per_line = max(max_chars_per_line, len(line))
    
    # حساب الارتفاع بناءً على حجم الخط وعدد الأسطر
    base_height = font_size + 8  # مساحة أساسية
    line_height = font_size + 4  # ارتفاع كل سطر إضافي
    
    # إضافة مساحة إضافية للأسطر الطويلة جداً
    if max_chars_per_line > 50:
        line_height += 2
    
    total_height = base_height + (max_lines - 1) * line_height
    
    # حد أدنى وأقصى معقول
    return max(20, min(total_height, 200))

def create_simple_excel_with_formatting(data_entries, filename):
    """
    إنشاء ملف Excel بسيط مع تنسيق جيد (بدون صور)
    
    Args:
        data_entries: بيانات الإدخالات
        filename: اسم الملف
    
    Returns:
        str: مسار الملف المؤقت
    """
    try:
        # التحقق من وجود البيانات
        if not data_entries:
            print("لا توجد بيانات للتصدير")
            return None
            
        # تحويل البيانات إلى DataFrame
        df_data = []
        for entry in data_entries:
            try:
                # التأكد من أن entry هو tuple أو list وله العدد المطلوب من العناصر
                if len(entry) < 12:
                    print(f"تحذير: السجل لا يحتوي على جميع الحقول المطلوبة: {len(entry)} حقل")
                    continue
                    
                df_data.append({
                    'Employee Name': str(entry[1]) if entry[1] else 'N/A',
                    'Employee Code': str(entry[2]) if entry[2] else 'N/A',
                    'Branch': str(entry[3]) if entry[3] else 'N/A',
                    'Shop Code': str(entry[4]) if entry[4] else 'N/A',
                    'Model': str(entry[5]) if entry[5] else 'N/A',
                    'Display Type': str(entry[6]) if entry[6] else 'N/A',
                    'Selected Materials': str(entry[7]).replace(',', '\n') if entry[7] else 'None',
                    'Missing Materials': str(entry[8]).replace(',', '\n') if entry[8] else 'None',
                    'Comments': str(entry[11]) if len(entry) > 11 and entry[11] else 'No comment',
                    'Date': str(entry[10]) if len(entry) > 10 and entry[10] else 'N/A'
                })
            except Exception as e:
                print(f"خطأ في معالجة السجل: {e}")
                continue
        
        if not df_data:
            print("لا توجد بيانات صالحة للتصدير")
            return None
            
        df = pd.DataFrame(df_data)
        
        # حفظ مع تنسيق
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='POP Materials Data', index=False)
            
            # تحسين التنسيق
            worksheet = writer.sheets['POP Materials Data']
            
            # استيراد الأنماط المطلوبة
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # تعريف الحدود
            thin_border = Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
            
            # تنسيق العناوين المحسن
            for cell in worksheet[1]:
                cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # تعيين ارتفاع صف العناوين
            worksheet.row_dimensions[1].height = 30
            
            # تنسيق البيانات وتوسيع الخلايا تلقائياً
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                # حساب ارتفاع الصف بناءً على المحتوى (محسن)
                max_lines = 1
                max_content_length = 0
                
                for cell in row:
                    # تطبيق التنسيق العام المحسن
                    cell.font = Font(name='Calibri', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = thin_border
                    
                    # حساب عدد الأسطر في الخلية بدقة أكبر
                    if cell.value and isinstance(cell.value, str):
                        # حساب الأسطر الفعلية
                        lines_count = cell.value.count('\n') + 1
                        max_lines = max(max_lines, lines_count)
                        
                        # حساب طول المحتوى لتقدير الحاجة لأسطر إضافية
                        content_length = len(cell.value)
                        max_content_length = max(max_content_length, content_length)
                
                # حساب ارتفاع الصف المطلوب (محسن)
                # الحد الأدنى: 20 نقطة
                # لكل سطر إضافي: 15 نقطة
                # إضافة مساحة إضافية للمحتوى الطويل
                base_height = 20
                line_height = 15
                content_bonus = min(10, max_content_length // 50)  # مساحة إضافية للمحتوى الطويل
                
                calculated_height = base_height + (max_lines - 1) * line_height + content_bonus
                
                # تعيين ارتفاع الصف (مع حد أدنى وأقصى معقول)
                final_height = max(25, min(calculated_height, 150))
                worksheet.row_dimensions[row_num].height = final_height
                
                # معلومات تشخيصية للصفوف التي تحتوي على محتوى كثير (تم إزالة print للأداء)
                # if max_lines > 3 or max_content_length > 100:
                #     pass  # تم إزالة print statement
            
            # تعديل عرض الأعمدة بذكاء أكبر
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            # حساب أطول سطر في الخلية
                            if isinstance(cell.value, str) and '\n' in cell.value:
                                lines = cell.value.split('\n')
                                cell_max_length = max(len(line) for line in lines)
                            else:
                                cell_max_length = len(str(cell.value))
                            
                            if cell_max_length > max_length:
                                max_length = cell_max_length
                    except:
                        pass
                
                # تعيين عرض العمود (مع حد أدنى وأقصى)
                adjusted_width = max(10, min(max_length + 3, 60))
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # إضافة Report Summary للملف البسيط
            add_report_summary_to_simple_excel(worksheet, data_entries)
        
        return temp_path
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel البسيط: {e}")
        return None

def add_report_summary_to_simple_excel(worksheet, data_entries):
    """
    إضافة ملخص التقرير للملف البسيط
    
    Args:
        worksheet: ورقة العمل
        data_entries: بيانات الإدخالات
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # تحديد الصف التالي بعد البيانات
        last_row = worksheet.max_row
        summary_start_row = last_row + 3
        
        # ألوان التنسيق
        header_color = '366092'
        
        # عنوان الملخص
        summary_title_cell = worksheet.cell(row=summary_start_row, column=1, value="Report Summary")
        summary_title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        summary_title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        summary_title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # دمج الخلايا للعنوان (عبر جميع الأعمدة)
        max_col = worksheet.max_column
        worksheet.merge_cells(f'A{summary_start_row}:{chr(64 + max_col)}{summary_start_row}')
        
        # حساب إحصائيات التقرير
        total_entries = len(data_entries)
        total_images = sum(len(entry[9].split(',')) if entry[9] else 0 for entry in data_entries)
        unique_employees = len(set(entry[2] for entry in data_entries))
        unique_branches = len(set(entry[3] for entry in data_entries))
        
        # بيانات الملخص مع الوقت المحلي الصحيح
        summary_data = [
            f"Total Entries: {total_entries}",
            f"Total Images: {total_images}",
            f"Unique Employees: {unique_employees}",
            f"Unique Branches: {unique_branches}",
            f"Report Generated: {get_local_time_string()}"  # استخدام الوقت المحلي الصحيح
        ]
        
        # إضافة بيانات الملخص
        for i, summary_text in enumerate(summary_data):
            row_num = summary_start_row + 1 + i
            cell = worksheet.cell(row=row_num, column=1, value=summary_text)
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # دمج الخلايا لكل سطر ملخص
            worksheet.merge_cells(f'A{row_num}:{chr(64 + max_col)}{row_num}')
        

        
    except Exception as e:
        print(f"⚠️ خطأ في إضافة Report Summary: {e}")
